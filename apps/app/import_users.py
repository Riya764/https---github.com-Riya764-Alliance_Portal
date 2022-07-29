'''
Imports Utility
'''
import csv
import logging
from django.db import IntegrityError
from django.http import HttpResponse
from django.contrib.auth.models import Group
from django.core.files.storage import FileSystemStorage
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook

from hul.choices import OrderDay
from hul.constants import (DEFAULT_PASSWORD, DEFAULT_COUNTRY,
                           FILE_TYPE_CSV, FILE_TYPE_XLSX)
from app.constants import (ALLIANCE_GROUP, DISTRIBUTER_GROUP,
                           REGIONAL_SALES_PROMOTER, SHAKTI_ENTERPRENEUR)
from app.models import (User, UserAddress, AlliancePartner,
                        RegionalSalesPromoter, ShaktiEntrepreneur,
                        RegionalDistributor)

from localization.models import State, Country
from product.import_products import ImportProduct


class ImportProcess(object):
    '''
    Process User Import File..
    '''
    @staticmethod
    def store_import_file(user_file):
        '''
        Save Uploaded File and return file path
        '''
        uploaded_file = FileSystemStorage()
        filename = uploaded_file.save(user_file.name, user_file)
        return 'run' + uploaded_file.base_url + filename

    @staticmethod
    def sample_export_csv(custom_fields, file_name):
        '''Sample Import CSV'''
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="' + \
            file_name + '"'
        writer = csv.writer(response)

        fields = []
        for field in custom_fields.keys():
            fields.append(custom_fields[field]['label'])
        writer.writerow(fields)
        return response

    @staticmethod
    def sample_export_xlsx(custom_fields, file_name):
        '''Sample Import CSV'''
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="' + \
            file_name + '"'

        workbook = Workbook()
        worksheet = workbook.active

        row = 1
        col = 1
        worksheet.append([row])
        for field in custom_fields.keys():
            worksheet.cell(column=col, row=row, value="{0}".format(
                custom_fields[field]['label']))
            col += 1

        workbook.save(response)
        return response

    @classmethod
    def save_obj(cls, worksheet, file_type, group_type, country, user_fields):
        '''
        Save User Object
        '''
        import_data = ImportData()
        impro = ImportProduct()
        error_list = []
        total_created = 0
        total_updated = 0
        line_num = 1

        reader = worksheet
        common_obj = {}
        common_obj['file_type'] = file_type
        common_obj['group_type'] = group_type
        common_obj['country'] = country
        total_updated = 0

        if file_type == FILE_TYPE_XLSX:
            reader = worksheet.iter_rows(min_row=2)

        for row in reader:
            try:
                line_num = import_data.get_line_num(row, reader, file_type)
                if impro.get_value(row[1], file_type) not in ['', None]:
                    email = impro.get_value(row[1], file_type).strip()
                    import_data.recheck_email(email)
                state = import_data.get_state(
                    impro.get_value(row[7], file_type))

                user_obj = import_data.get_user_obj(row, file_type, group_type)

                common_obj['total_created'] = total_created
                common_obj['total_updated'] = total_updated
                common_obj['state'] = state

                total_created, total_updated = import_data.save_user_obj(
                    user_obj, row, common_obj)
            except IntegrityError as error:
                message = ErrorMessages.prepare_unique_message(
                    error.message, user_fields)
                error_list.append(
                    "Error on line no. %s %s" % (line_num, message))
            except StandardError as error:
                logging.exception(error.message)
                error_list.append(
                    "Error on line no. %s %s" % (line_num, error.message))
        return total_created, total_updated, error_list

    @classmethod
    def save_import_file(cls, request, import_file, group_type, user_fields):
        '''
        Save Uploaded File
        '''
        error_list = []
        impro = ImportProduct()
        country = Country.objects.filter(
            name__icontains=DEFAULT_COUNTRY).first()

        if import_file.split('.')[1] == FILE_TYPE_CSV:
            with open(import_file, 'r') as userfile:
                next(userfile)
                worksheet = csv.reader(userfile)
                total_created, total_updated, error_list = cls.save_obj(
                    worksheet, FILE_TYPE_CSV, group_type, country, user_fields)

        elif import_file.split('.')[1] == FILE_TYPE_XLSX:
            workbook = load_workbook(filename=import_file, read_only=True)
            worksheet = workbook.active
            total_created, total_updated, error_list = cls.save_obj(
                worksheet, FILE_TYPE_XLSX, group_type, country, user_fields)
        impro.show_messages(request, int(total_created), int(total_updated))

        return error_list


class ImportData(object):
    '''
    Import Data
    '''

    def save_user_obj(self, user_obj, row, common_obj):
        '''
        Create/Update User Obj
        '''
        impro = ImportProduct()

        file_type = common_obj['file_type']
        group_type = common_obj['group_type']
        state = common_obj['state']
        country = common_obj['country']
        total_created = common_obj['total_created']
        total_updated = common_obj['total_updated']

        if user_obj:
            user_obj.user.name = impro.get_value(row[0], file_type)
            user_obj.user.contact_number = impro.get_value(
                row[3], file_type)
            user_obj.address.address_line1 = impro.get_value(
                row[4], file_type)
            user_obj.address.address_line2 = impro.get_value(
                row[5], file_type)
            user_obj.address.address_line3 = impro.get_value(
                row[6], file_type)
            user_obj.address.state = state
            user_obj.address.city = impro.get_value(row[8], file_type)
            user_obj.address.post_code = impro.get_value(
                row[9], file_type)
            password = impro.get_value(row[10], file_type)

            if password not in ['', None]:
                user_obj.user.set_password(password)

            self.update_imported_user(
                row, group_type, file_type, user_obj)

            user_obj.user.save()
            user_obj.address.save()
            user_obj.save()

            total_updated += 1

        else:
            '''
            insert user and other details
            '''

            user, created = self.add_user(
                impro.get_value(row[2], file_type),
                impro.get_value(row[10], file_type),
                impro.get_value(row[0], file_type),
                impro.get_value(row[1], file_type),
                impro.get_value(row[3], file_type),
                impro.get_value(row[11], file_type),
                group_type)
            if not created:
                raise StandardError(
                    ' Column USERNAME and EMAIL already exists')

            address = self.add_address(impro.get_value(row[4], file_type),
                                       impro.get_value(row[5], file_type),
                                       impro.get_value(row[6], file_type),
                                       impro.get_value(row[8], file_type),
                                       state, country, impro.get_value(row[9], file_type))

            response = self.add_imported_user(
                row, user, address, group_type, file_type)

            if response:
                total_created += 1
            else:
                total_updated += 1
        return total_created, total_updated

    @staticmethod
    def get_line_num(row, reader, file_type):
        '''
        return processed line num
        '''
        if file_type == FILE_TYPE_XLSX:
            line_num = row[0].row - 1
        elif file_type == FILE_TYPE_CSV:
            line_num = reader.line_num
        return line_num

    @staticmethod
    def recheck_email(email):
        '''Validate Email'''
        try:
            validate_email(email)
            return '', ''
        except ValidationError as error:
            raise StandardError(error.message)

    @staticmethod
    def get_state(state):
        '''Return State Object '''
        if state:
            try:
                return State.objects.filter(name__iexact=state).get()
            except State.DoesNotExist:
                raise StandardError('State is not defined')
        else:
            raise StandardError('State is not defined')

    @staticmethod
    def add_user(username, password, name, email, contact_number, code, group_type):
        '''' add user '''

        if email not in ['', None]:
            email = email.strip()
        else:
            email = None
        if not username:
            username = name.strip(',-() ').replace(" ", ".").lower()
        else:
            username = username.strip(',-() ').replace(" ", ".").lower()

        if group_type == SHAKTI_ENTERPRENEUR:
            username = code

        user_obj = User.objects.get_or_create(username__iexact=username,
                                              email__iexact=email,
                                              defaults={
                                                  'name': name,
                                                  'username': username,
                                                  'email': email,
                                                  'contact_number': contact_number})

        if password is None or len(str(password)) < 2:
            password = DEFAULT_PASSWORD

        user_obj[0].set_password(password)
        if group_type in [ALLIANCE_GROUP, DISTRIBUTER_GROUP]:
            user_obj[0].is_staff = True
        user_obj[0].save()

        if user_obj[1]:
            group = Group.objects.filter(name__iexact=group_type).first()
            if group:
                group.user_set.add(user_obj[0])
        return user_obj

    def add_imported_user(self, row, user, address, group_type, file_type):
        ''' Add csv users '''
        response = False
        impro = ImportProduct()
        if group_type == ALLIANCE_GROUP:
            code = str(impro.get_value(row[11], file_type))
            result = self.add_alliance(
                user, address, code)

            response = result[1]

        elif group_type == DISTRIBUTER_GROUP:
            result = self.add_regional_distributor(
                row, user, address, file_type)
            response = result

        elif group_type == REGIONAL_SALES_PROMOTER:

            result = self.add_sales_promoter(row,
                                             user, address, file_type)
            response = result[1]

        elif group_type == SHAKTI_ENTERPRENEUR:
            result = self.add_shakti_entrepreneur(row,
                                                  user, address,
                                                  file_type)
            response = result[1]

        return response

    @staticmethod
    def update_imported_user(row, group_type, file_type, user_obj):
        ''' update csv users '''
        impro = ImportProduct()

        if group_type == DISTRIBUTER_GROUP:

            order_day = [x[0] for x in OrderDay.STATUS
                         if impro.get_value(row[13], file_type).upper() == x[1]][0]
            user_obj.order_day = order_day
            user_obj.min_order = impro.get_value(row[14], file_type)
            user_obj.ap_dist_channel = impro.get_value(row[15], file_type)
            user_obj.ap_division = impro.get_value(row[16], file_type)
            user_obj.ap_plant = impro.get_value(row[17], file_type)

            alliance_partner = AlliancePartner.objects.\
                filter(code__iexact=str(
                    impro.get_value(row[12], file_type))).first()
            if alliance_partner:
                user_obj.alliance_partner.add(alliance_partner)
            else:
                raise StandardError(
                    "Column ALLIANCE PARTNER CODE does not exists.")

        elif group_type == REGIONAL_SALES_PROMOTER:
            user_obj.employee_number = impro.get_value(row[13], file_type)
            regional_distributor = RegionalDistributor.objects.\
                filter(code__iexact=str(
                    impro.get_value(row[12], file_type))).first()
            if regional_distributor:
                user_obj.regional_distributor = regional_distributor
            else:
                raise StandardError(
                    "Column REDISTRIBUTION STOCKIST CODE does not exists.")

        elif group_type == SHAKTI_ENTERPRENEUR:
            regional_sales = RegionalSalesPromoter.objects.\
                filter(rsp_id__iexact=str(
                    impro.get_value(row[12], file_type))).first()

            if regional_sales:
                user_obj.regional_sales = regional_sales
                user_obj.code = impro.get_value(row[11], file_type)
                # added hul code as Shakti username
                user_obj.user.username = impro.get_value(row[11], file_type)
                user_obj.beat_name = impro.get_value(row[13], file_type)
                user_obj.is_active = impro.get_value(row[16], file_type) in [
                    'True', 'true', 'TRUE']

                try:
                    order_day = [x[0] for x in OrderDay.STATUS
                                 if impro.get_value(row[14], file_type).upper() == x[1]][0]
                    user_obj.order_day = order_day
                except IndexError:
                    raise StandardError(" Column ORDER DAY not valid.")
            else:
                raise StandardError(" Column RSP ID does not exists.")

    @staticmethod
    def get_user_obj(row, file_type, group_type):
        ''' Add csv users '''
        response = None
        impro = ImportProduct()
        if group_type == ALLIANCE_GROUP:
            response = AlliancePartner.objects.filter(
                user__username__iexact=impro.get_value(row[2], file_type),
                user__email__iexact=impro.get_value(row[1], file_type),
                code=impro.get_value(row[11], file_type)).first()

        elif group_type == DISTRIBUTER_GROUP:
            response = RegionalDistributor.objects.filter(
                user__username__iexact=impro.get_value(row[2], file_type),
                user__email__iexact=impro.get_value(row[1], file_type),
                code=impro.get_value(row[11], file_type)).first()

        elif group_type == REGIONAL_SALES_PROMOTER:
            email = None
            if impro.get_value(row[1], file_type) != '':
                email = impro.get_value(row[1], file_type)
            response = RegionalSalesPromoter.objects.filter(
                user__username__iexact=impro.get_value(row[2], file_type),
                user__email__iexact=email,
                rsp_id=impro.get_value(row[11], file_type)).first()

        elif group_type == SHAKTI_ENTERPRENEUR:
            response = ShaktiEntrepreneur.objects.filter(user__username__iexact=impro.get_value(row[11], file_type),
                                                         code=impro.get_value(row[11], file_type)).first()

        return response

    @staticmethod
    def add_address(address_line1, address_line2,
                    address_line3, city, state, country, post_code):
        '''
        Add address
        '''
        return UserAddress.objects.create(
            address_line1=address_line1,
            address_line2=address_line2,
            address_line3=address_line3,
            city=city,
            state=state,
            country=country,
            post_code=post_code
        )

    @staticmethod
    def add_alliance(user, address, code):
        '''
        Add Alliance User
        '''
        return AlliancePartner.objects.\
            update_or_create(user=user.id,
                             defaults={
                                 'user': user,
                                 'address': address,
                                 'code': code,

                             })

    @staticmethod
    def add_sales_promoter(row, user, address, file_type):
        '''
        Add RSP User
        '''
        impro = ImportProduct()
        regional_distributor = RegionalDistributor.objects.\
            filter(code__iexact=str(
                impro.get_value(row[12], file_type))).first()
        if regional_distributor:
            rsp_id = str(impro.get_value(row[11], file_type))
            return RegionalSalesPromoter.objects.\
                update_or_create(regional_distributor=regional_distributor,
                                 rsp_id__iexact=rsp_id,
                                 defaults={
                                     'regional_distributor': regional_distributor,
                                     'user': user,
                                     'address': address,
                                     'rsp_id': rsp_id,
                                     'employee_number': impro.get_value(row[13], file_type)
                                 })
        else:
            User.objects.get(id=user.id).delete()
            UserAddress.objects.get(id=address.id).delete()
            raise StandardError(
                "Column REDISTRIBUTION STOCKIST CODE does not exists.")

    @staticmethod
    def add_shakti_entrepreneur(row, user, address, file_type):
        '''
        Add Shakti Entrepreneur User
        '''
        impro = ImportProduct()
        regional_sales = RegionalSalesPromoter.objects.\
            filter(rsp_id__iexact=str(
                impro.get_value(row[12], file_type))).first()

        if regional_sales:
            code = str(impro.get_value(row[11], file_type))
            try:
                order_day = [x[0] for x in OrderDay.STATUS
                             if impro.get_value(row[14], file_type).upper() == x[1]][0]
                return ShaktiEntrepreneur.\
                    objects.update_or_create(regional_sales=regional_sales,
                                             code__iexact=code, defaults={
                                                 'regional_sales': regional_sales,
                                                 'user': user,
                                                 'address': address,
                                                 'code': code,
                                                 'beat_name': impro.get_value(row[13], file_type),
                                                 'order_day': order_day,
                                                 'min_order': impro.get_value(row[15], file_type),
                                                 'is_active': impro.get_value(row[16], file_type),
                                             })
            except IndexError:
                User.objects.get(id=user.id).delete()
                UserAddress.objects.get(id=address.id).delete()
                raise StandardError(" Column ORDER DAY does not exists.")

        else:
            User.objects.get(id=user.id).delete()
            UserAddress.objects.get(id=address.id).delete()
            raise StandardError(" Column RSP ID does not exists.")

    @staticmethod
    def add_regional_distributor(row, user, address, file_type):
        '''
        Add Regional Distributor User
        '''
        alliance_partner = AlliancePartner.objects.filter(
            code__iexact=str(ImportProduct().get_value(row[12], file_type))).first()
        if alliance_partner:
            code = str(ImportProduct().get_value(row[11], file_type))
            min_order = float(ImportProduct().get_value(row[14], file_type))

            ap_dist_channel = ImportProduct().get_value(
                row[15], file_type) or ''
            ap_division = ImportProduct().get_value(
                row[16], file_type) or ''
            ap_plant = ImportProduct().get_value(row[17], file_type) or ''
            gst_code = ImportProduct().get_value(row[18], file_type) or ''
            hul_code = ImportProduct().get_value(row[19], file_type) or ''
            regional_distributor, created = RegionalDistributor.objects.update_or_create(
                code=code,
                defaults={'min_order': min_order,
                          'ap_dist_channel': ap_dist_channel,
                          'ap_division': ap_division,
                          'ap_plant': ap_plant,
                          'gst_code': gst_code,
                          'hul_code': hul_code
                          },)

            regional_distributor.user = user
            regional_distributor.address = address
            regional_distributor.save()
            regional_distributor.alliance_partner.add(alliance_partner)
            return regional_distributor
        else:
            User.objects.get(id=user.id).delete()
            UserAddress.objects.get(id=address.id).delete()
            raise StandardError(
                "Column ALLIANCE PARTNER CODE does not exists.")


class ErrorMessages(object):
    ''' handle Error Messages '''
    @staticmethod
    def prepare_unique_message(error_message, columns):
        ''' Prepare unique messages '''
        message = ''
        keys = columns.keys()
        for key in keys:
            if key in error_message:
                message = 'Column %s already exists' % (columns[key]['label'])

        return message
