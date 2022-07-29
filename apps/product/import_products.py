'''
Imports Utility
'''
import csv
import logging
from django.contrib import messages
from openpyxl.reader.excel import load_workbook
from hul.constants import FILE_TYPE_CSV, FILE_TYPE_XLSX
from product.models import Product, Category
from app.models import AlliancePartner
from localization.models import MeasurementUnit


class ImportProduct(object):
    '''
    Process User Import File..
    '''
    @classmethod
    def save_obj(cls, request, reader, file_type):
        '''
        Save User Object
        '''
        total_updated = 0
        total_created = 0
        impro = ImportProduct()
        error_list = []
        line_num = 1

        for row in reader:
            line_num = impro.get_line_num(row, reader, file_type)
            try:

                category = Category.objects.get_or_create(
                    name__iexact=str(cls.get_value(row[0], file_type)),
                    defaults={'name': str(cls.get_value(row[0], file_type))})

                brand = AlliancePartner.objects.filter(
                    code__iexact=str(cls.get_value(row[1], file_type))).first()

                unit = MeasurementUnit.objects.get_or_create(
                    unit__iexact=str(cls.get_value(row[5], file_type)),
                    defaults={'unit': str(cls.get_value(row[5], file_type))})
                result = Product.objects.update_or_create(
                    category=category[0],
                    brand=brand,
                    basepack_name__iexact=str(
                        cls.get_value(row[2], file_type)),
                    basepack_code__iexact=str(
                        cls.get_value(row[3], file_type)),
                    basepack_size__iexact=str(
                        cls.get_value(row[4], file_type)),
                    defaults={'category': category[0],
                              'brand': brand,
                              'partner_code': brand.code,
                              'basepack_name':
                              cls.get_value(
                        row[2], file_type),
                        'basepack_code':
                        cls.get_value(
                        row[3], file_type),
                        'basepack_size':
                        cls.get_value(
                        row[4], file_type),
                        'unit': unit[0],
                        'expiry_day':
                        cls.get_value(
                            row[6], file_type) or None,
                        'cld_configurations':
                        cls.get_value(
                        row[7], file_type),
                        'mrp':
                        float(cls.get_value(
                            row[8], file_type)),
                        'base_rate':
                        float(
                        cls.get_value(row[9],
                                      file_type)),
                        'cgst': float(cls.get_value(row[10], file_type) or 0),
                        'sgst': float(cls.get_value(row[11], file_type) or 0),
                        'igst': float(cls.get_value(row[12], file_type) or 0),
                        'hsn_code': cls.get_value(row[13], file_type)
                    })

                if not result[1]:
                    total_updated = total_updated + 1
                elif result[1]:
                    total_created = total_created + 1
            except StandardError as error:
                raise StandardError("Error on line %s - %s" %
                                    (line_num, error.message))

        cls.show_messages(request, total_created, total_updated)

    @classmethod
    def save_import_file(cls, request, import_file):
        '''
        Save Uploaded File
        '''
        if import_file.split('.')[1] == FILE_TYPE_CSV:
            with open(import_file, 'r') as userfile:
                next(userfile)
                reader = csv.reader(userfile)
                try:
                    cls.save_obj(
                        request, reader, FILE_TYPE_CSV)
                except StandardError as error:
                    logging.exception(error.message)
                    messages.error(
                        request, 'No records has been created/updated.')

        elif import_file.split('.')[1] == FILE_TYPE_XLSX:
            workbook = load_workbook(filename=import_file, read_only=True)
            worksheet = workbook.active
            reader = worksheet.iter_rows(min_row=2)
            try:
                cls.save_obj(request, reader, FILE_TYPE_XLSX)
            except StandardError as error:
                raise StandardError(error.message)
                messages.error(request, 'No records has been created/updated.')

    @staticmethod
    def get_value(col, file_type):
        '''
        get value of a cell according to file type
        '''
        if file_type == FILE_TYPE_XLSX:
            col = col.value

        if isinstance(col, str):
            col = col.strip()

        if isinstance(col, long):
            col = int(col)

        return col

    @staticmethod
    def show_messages(request, total_created, total_updated):
        '''
        display messages for records updated/created
        '''
        final_text = ''
        if int(total_updated) == 0 and int(total_created) == 0:
            messages.error(request, 'No records has been created/updated.')
        else:
            if total_created:
                final_text = final_text + \
                    "%s row(s) imported" % (total_created)
            if total_updated:
                final_text = final_text + "%s row(s) updated" % (total_updated)

            final_text = final_text + " successfully. "
            messages.success(request, final_text)

    @staticmethod
    def get_line_num(row, reader, file_type):
        '''
        return processed line num
        '''
        if file_type == FILE_TYPE_XLSX:
            line_num = row[0].row - 1
        elif file_type == FILE_TYPE_CSV:
            line_num = reader.line_num
        return line_num
